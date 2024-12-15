import json
import datetime
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_datetime
from openpyxl import Workbook
from django.http import HttpResponse
from .models import Robot


# Отключаем проверку CSRF для упрощения (можно включить позже)
@csrf_exempt
def add_robot(request):
    if request.method == 'POST':
        try:
            # Получаем данные из тела запроса
            data = json.loads(request.body)
            model = data.get('model')
            version = data.get('version')
            created = data.get('created')

            # Валидация модели
            valid_models = ['R2', '13', 'X5']  # Список допустимых моделей
            if model not in valid_models:
                return JsonResponse({"error": f"Недопустимая модель: {model}"},
                                    status=400,
                                    json_dumps_params={'ensure_ascii': False})

            # Валидация версии
            valid_versions = ['D2', 'XS', 'LT']  # Список допустимых версий
            if version not in valid_versions:
                return JsonResponse({"error": f"Недопустимая версия: {version}"},
                                    status=400,
                                    json_dumps_params={'ensure_ascii': False})

            # Валидация даты
            created_datetime = parse_datetime(created)
            if not created_datetime:
                return JsonResponse({"error": "Некорректный формат даты. Используйте YYYY-MM-DD HH:MM:SS"},
                                    status=400,
                                    json_dumps_params={'ensure_ascii': False})

            # Генерация серийного номера
            serial = f"{model}-{version}-{created_datetime.strftime('%Y%m%d%H%M%S')}"

            # Сохранение записи в базе данных
            Robot.objects.create(serial=serial, model=model, version=version, created=created_datetime)

            return JsonResponse(
                {"message": "Робот успешно добавлен", "serial": serial},
                status=201,
                json_dumps_params={'ensure_ascii': False}  # Отключаем ASCII-кодирование
            )

        except json.JSONDecodeError:
            return JsonResponse({"error": "Некорректный формат JSON"},
                                status=400,
                                json_dumps_params={'ensure_ascii': False})

    return JsonResponse({"error": "Метод не поддерживается. Используйте POST"},
                        status=405,
                        json_dumps_params={'ensure_ascii': False})

def download_report(request):
    # Определяем диапазон дат (последняя неделя)
    today = datetime.date.today()
    week_ago = today - datetime.timedelta(days=7)

    # Фильтруем роботов по дате создания
    robots = Robot.objects.filter(created__date__gte=week_ago)

    # Группируем данные по моделям и версиям
    data = {}
    for robot in robots:
        model = robot.model
        version = robot.version
        if model not in data:
            data[model] = {}
        if version not in data[model]:
            data[model][version] = 0
        data[model][version] += 1

    # Создаём Excel-файл
    wb = Workbook()

    for model, versions in data.items():
        # Создаём новый лист для каждой модели
        sheet = wb.create_sheet(title=model)
        sheet.append(["Модель", "Версия", "Количество за неделю"])  # Заголовки

        for version, count in versions.items():
            # Добавляем строки для каждой версии
            sheet.append([model, version, count])

    # Удаляем стандартный лист, если он есть
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb['Sheet'])

    # Возвращаем файл как ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="robots_report.xlsx"'
    wb.save(response)

    return response

